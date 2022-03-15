import openpyxl
from sys import argv
from PyQt5 import QtGui
from PyQt5 import QtCore
from PyQt5 import QtWidgets


class Window(QtWidgets.QWidget):
    """Создание графического интерфейса"""
    def __init__(self):
        """Инициализация графического интерфейса"""
        super(Window, self).__init__()
        self.set_settings_main_window()  # Настройки главного окна
        self.style_sheet = "font: 14pt \"Times New Roman\";"  # Размер и шрифт текста

        self.count = 1  # Счетчик для добавления строк
        self.flag = False  # Флаг для активации сохранения

        self.combo_box = QtWidgets.QComboBox()  # Создание комбобокса
        self.set_settings_combo_boxes()  # Настройки комбобокса
        self.line_edit = QtWidgets.QLineEdit()  # Создание поля ввода
        self.set_settings_line_edits()  # Настройки поля ввода

        h_layout_1 = QtWidgets.QHBoxLayout()  # Создание горизонтального лейаута
        for i in [self.combo_box, self.line_edit]:
            h_layout_1.addWidget(i)  # Добавление виджетов

        self.search_button = QtWidgets.QPushButton()  # Создание кнопки "Поиск"
        self.add_button = QtWidgets.QPushButton()  # Создание кнопки "Добавить"
        self.del_button = QtWidgets.QPushButton()  # Создание кнопки "Удалить"
        self.save_button = QtWidgets.QPushButton()  # Создание кнопки "Сохранить"
        self.set_settings_buttons()  # Настрока кнопок

        h_layout_2 = QtWidgets.QHBoxLayout()  # Создание горизонтального лейаута
        for i in [self.search_button, self.save_button, self.add_button, self.del_button]:
            h_layout_2.addWidget(i)  # Добавление виджетов

        h_layout_3 = QtWidgets.QHBoxLayout()  # Создание горизонтального лейаута
        for i in [h_layout_1, h_layout_2]:
            h_layout_3.addLayout(i)  # Добавление лейаутов

        self.table_widget = QtWidgets.QTableWidget()  # Создание таблицы
        self.header_list = ["Улица", "Дом", "Место", "TKD", "IP", "Примечание"]  # Названия хедера таблицы
        self.width_header = [250, 100, 160, 125, 125, 395]
        self.set_settings_table_widget()  # Настройка таблицы

        v_layout = QtWidgets.QVBoxLayout()  # Создание вертикального лейаута
        v_layout.addLayout(h_layout_3)  # Добавление лейатута
        v_layout.addWidget(self.table_widget)  # Добавление виджета

        self.setLayout(v_layout)  # Отображение вертикального лейаута

    def set_settings_main_window(self):
        """Найстройка главного окна"""
        self.resize(1200, 700)
        self.setWindowTitle("NVBS")
        self.setWindowIcon(QtGui.QIcon("Icons/main.ico"))

    def set_settings_combo_boxes(self):
        """Настройка поля Список"""
        self.combo_box.setMinimumSize(30, 30)
        self.combo_box.setStyleSheet(self.style_sheet)
        self.combo_box.setToolTip("Выберете раздел поиска")

        for i in ["Улица", "TKD", "IP"]:
            self.combo_box.addItem(i)

    def set_settings_line_edits(self):
        """Настройка поля Ввода"""
        self.line_edit.setMinimumSize(QtCore.QSize(0, 30))
        self.line_edit.setStyleSheet(self.style_sheet)
        self.line_edit.setToolTip("Введите значение")
        self.line_edit.returnPressed.connect(self.search_table)

    def set_settings_buttons(self):
        """Настройка кнопок"""
        self.search_button.setIcon(QtGui.QIcon("Icons/search.ico"))
        self.search_button.setToolTip("Поиск")
        self.search_button.clicked.connect(self.search_table)

        self.add_button.setIcon(QtGui.QIcon("Icons/add.ico"))
        self.add_button.setToolTip("Добавить строку")
        self.add_button.clicked.connect(self.add_table)

        self.del_button.setIcon(QtGui.QIcon("Icons/del.ico"))
        self.del_button.setToolTip("Удалить строку")
        self.del_button.clicked.connect(self.del_table_in_db)

        self.save_button.setIcon(QtGui.QIcon("Icons/save.ico"))
        self.save_button.setToolTip("Сохранить")
        self.save_button.clicked.connect(self.save_table_in_db)

        for i in [self.search_button, self.add_button, self.del_button, self.save_button]:
            i.setMinimumSize(33, 32)
            i.setStyleSheet(self.style_sheet)
            i.setToolTipDuration(3000)

    def set_settings_table_widget(self):
        """Настройка таблицы"""
        self.table_widget.setColumnCount(len(self.header_list))
        self.table_widget.setHorizontalHeaderLabels(self.header_list)
        self.table_widget.setStyleSheet(self.style_sheet)
        self.table_widget.verticalHeader().setVisible(False)
        self.table_widget.setSortingEnabled(True)
        self.table_widget.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_widget.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.SingleSelection)

        for i, j in zip(self.header_list, self.width_header):
            self.table_widget.setColumnWidth(self.header_list.index(i), j)

    def search_table(self):
        """Поиск по БД"""
        column_index = self.header_list.index(self.combo_box.currentText())
        string_index = 1
        for i in range(2, sheet.max_row + 1):
            if self.line_edit.text().lower() in str(sheet[i][column_index].value).lower():
                self.table_widget.setRowCount(string_index)
                for j in range(len(self.header_list)):
                    text = str(sheet[i][j].value)
                    if text != "None":
                        self.table_widget.setItem(string_index - 1, j, QtWidgets.QTableWidgetItem(text))
                string_index += 1
            if string_index == 1:
                self.table_widget.setRowCount(0)
        self.count = 0
        self.flag = False

    def add_table(self):
        """Добавляет строки в таблицу"""
        if not self.count:
            self.table_widget.setRowCount(self.count)
            self.count += 1
            self.table_widget.setRowCount(self.count)
        self.table_widget.setRowCount(self.count)
        self.count += 1
        self.flag = True

    def save_table_in_db(self):
        """Сохранение добавленных данных в БД"""
        if self.flag:
            array = [[self.table_widget.item(j, i).text()
                      if self.table_widget.item(j, i) else ""
                      for i in range(len(self.header_list))]
                     for j in range(0, self.table_widget.rowCount())]

            for i in array:
                if i != [""] * 6:
                    sheet.append(i)
            book.save("DataBase.xlsx")

    def del_table_in_db(self):
        """Удаление строки из БД"""
        array = [self.table_widget.item(i.row(), i.column()).text()
                 if self.table_widget.item(i.row(), i.column())
                 else str(self.table_widget.item(i.row(), i.column()))
                 for i in self.table_widget.selectedIndexes()]

        for i in range(2, sheet.max_row + 1):
            count = 0
            if array[0] == str(sheet[i][0].value):
                count += 1
                for j in range(1, len(self.header_list)):
                    if array[j] == str(sheet[i][j].value):
                        count += 1
            if count == len(self.header_list):
                sheet.delete_rows(i)
                break
        book.save("DataBase.xlsx")


if __name__ == "__main__":
    book = openpyxl.load_workbook("DataBase.xlsx")
    sheet = book.active
    app = QtWidgets.QApplication(argv)
    window = Window()
    window.show()
    app.exec()

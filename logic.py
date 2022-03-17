from settings import Settings, QtWidgets
import openpyxl

book = openpyxl.load_workbook("DataBase.xlsx")
sheet = book.active


class Logic(Settings):
    """Найстрока логики графического интерфейса"""
    def __init__(self):
        super(Logic, self).__init__()

        self.count = 1  # Счетчик для добавления строк
        self.flag = False  # Флаг для активации сохранения

        self.line_edit.returnPressed.connect(self.search_table)
        self.search_button.clicked.connect(self.search_table)
        self.add_button.clicked.connect(self.add_table)
        self.del_button.clicked.connect(self.del_table_in_db)
        self.save_button.clicked.connect(self.save_table_in_db)

    def search_table(self):
        """Поиск по БД"""
        column_index = self.header_list.index(self.combo_box.currentText())
        string_index = 1
        for i in range(2, sheet.max_row + 1):
            if self.line_edit.text().lower() in str(sheet[i][column_index].value).lower():
                self.table_widget.setRowCount(string_index)
                for j in range(len(self.header_list)):
                    text = str(sheet[i][j].value)
                    if text != "None":
                        self.table_widget.setItem(string_index - 1, j, QtWidgets.QTableWidgetItem(text))
                string_index += 1
            if string_index == 1:
                self.table_widget.setRowCount(0)
        self.count = 0
        self.flag = False

    def add_table(self):
        """Добавляет строки в таблицу"""
        if not self.count:
            self.table_widget.setRowCount(self.count)
            self.count += 1
            self.table_widget.setRowCount(self.count)
        self.table_widget.setRowCount(self.count)
        self.count += 1
        self.flag = True

    def save_table_in_db(self):
        """Сохранение добавленных данных в БД"""
        if self.flag:
            array = [[self.table_widget.item(j, i).text()
                      if self.table_widget.item(j, i) else ""
                      for i in range(len(self.header_list))]
                     for j in range(0, self.table_widget.rowCount())]

            for i in array:
                if i != [""] * len(self.header_list):
                    sheet.append(i)
            book.save("DataBase.xlsx")

    def del_table_in_db(self):
        """Удаление строки из БД"""
        array = [self.table_widget.item(i.row(), i.column()).text()
                 if self.table_widget.item(i.row(), i.column())
                 else str(self.table_widget.item(i.row(), i.column()))
                 for i in self.table_widget.selectedIndexes()]

        for i in range(2, sheet.max_row + 1):
            count = 0
            if array[0] == str(sheet[i][0].value):
                count += 1
                for j in range(1, len(self.header_list)):
                    if array[j] == str(sheet[i][j].value):
                        count += 1
            if count == len(self.header_list):
                sheet.delete_rows(i)
                break
        book.save("DataBase.xlsx")
